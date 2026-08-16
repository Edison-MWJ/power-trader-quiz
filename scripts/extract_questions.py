#!/usr/bin/env python3
"""Extract exam question banks into a browser-friendly JS data file."""

from __future__ import annotations

import json
import re
import subprocess
from datetime import datetime
from pathlib import Path

from docx import Document
from openpyxl import load_workbook


SOURCE_DIR = Path(
    "/Users/higher/Library/CloudStorage/OneDrive-个人/02个人文档资料/01学习资料/"
    "交易员考试/交易员考试"
)
SOURCES = {
    "中级工": SOURCE_DIR / "中级+高级+技师" / "电力交易员（中级工）题库.xlsx",
    "高级工": SOURCE_DIR / "中级+高级+技师" / "电力交易员（高级工）题库.xlsx",
    "技师": SOURCE_DIR / "中级+高级+技师" / "电力交易员（技师）题库.xlsx",
}
PDF_SOURCE_DIRS = [
    SOURCE_DIR / "高级+技师试卷",
    SOURCE_DIR / "电力交易员" / "样卷2",
    SOURCE_DIR / "电力交易员" / "样卷3",
    SOURCE_DIR / "电力交易员-2026电子题库(1)",
]
DOCX_SOURCES = {
    "高级工": SOURCE_DIR / "电力交易员高级工题库-答案版).docx",
}
ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "questions.js"
DATA_DIR = OUTPUT.parent
INDEX = ROOT / "index.html"
SERVICE_WORKER = ROOT / "service-worker.js"
EXPLANATION_OVERRIDES = ROOT / "scripts" / "advanced_explanation_overrides.json"
CHUNK_SIZE = 180
LETTERS = "ABCDEFGHIJ"
JUDGMENT_OPTIONS = [{"label": "对", "text": "对"}, {"label": "错", "text": "错"}]
SCRIPT_BLOCK_RE = re.compile(
    r'\n  <script src="data/meta\.js"></script>\n'
    r"  <script>window\.QUESTION_PARTS = \[\];</script>\n"
    r'(?:  <script src="data/questions-\d{2}\.js"></script>\n)+'
    r"  <script>window\.QUESTION_BANK = \{ meta: window\.QUESTION_META, questions: window\.QUESTION_PARTS\.flat\(\) \};</script>"
)
APP_SHELL_RE = re.compile(r"const APP_SHELL = \[\n.*?\n\];", re.S)


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def compact(value: object) -> str:
    return re.sub(r"\s+", "", clean(value))


def canonical(value: object) -> str:
    text = compact(value).upper()
    table = str.maketrans(
        {
            "，": ",",
            "。": ".",
            "；": ";",
            "：": ":",
            "？": "?",
            "！": "!",
            "（": "(",
            "）": ")",
            "【": "[",
            "】": "]",
            "“": '"',
            "”": '"',
            "‘": "'",
            "’": "'",
            "．": ".",
            "、": ",",
            "－": "-",
            "—": "-",
            "–": "-",
            "％": "%",
        }
    )
    return re.sub(r"[。\.]+$", "", text.translate(table))


def normalize_answer(value: object, qtype: str) -> list[str]:
    text = compact(value)
    if qtype == "判断":
        return [
            text.replace("正确", "对")
            .replace("错误", "错")
            .replace("√", "对")
            .replace("×", "错")
            .replace("A", "对")
            .replace("B", "错")
        ]
    return [letter for letter in LETTERS if letter in text.upper()]


def question_scope(levels: list[str]) -> str:
    if len(levels) == 1:
        return f"{levels[0]}独有"
    return "+".join(levels)


def extract_xlsx_bank(level: str, source: Path) -> list[dict[str, object]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    questions = []

    for sheet in workbook.worksheets:
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            stem = clean(row[0] if len(row) > 0 else "")
            qtype = clean(row[1] if len(row) > 1 else "")
            answer = clean(row[2] if len(row) > 2 else "")
            if not stem or not qtype or not answer:
                continue

            options = []
            if qtype in {"单选", "多选"}:
                for idx, value in enumerate(row[3:13]):
                    label = LETTERS[idx]
                    text = clean(value)
                    if text:
                        options.append({"label": label, "text": text})
            elif qtype == "判断":
                options = [{"label": "对", "text": "对"}, {"label": "错", "text": "错"}]

            questions.append(
                {
                    "level": level,
                    "origin": source.name,
                    "id": f"{level}-{sheet.title}-{row_number}",
                    "type": qtype,
                    "stem": stem,
                    "options": options,
                    "answer": normalize_answer(answer, qtype),
                }
            )

    return questions


def extract_docx_bank(level: str, source: Path) -> list[dict[str, object]]:
    document = Document(source)
    questions: list[dict[str, object]] = []
    qtype = ""
    current: dict[str, object] | None = None

    def flush_current() -> None:
        nonlocal current
        if current and current.get("stem") and current.get("answer"):
            questions.append(current)
        current = None

    def append_explanation(text: str) -> None:
        if current is None:
            return
        explanation = clean(text)
        if not explanation:
            return
        existing = clean(current.get("explanation", ""))
        current["explanation"] = clean(f"{existing} {explanation}") if existing else explanation

    for paragraph in document.paragraphs:
        text = clean(paragraph.text)
        if not text:
            continue

        if text == "一、单选题":
            flush_current()
            qtype = "单选"
            continue
        if text == "二、多选题":
            flush_current()
            qtype = "多选"
            continue
        if text == "判断题":
            flush_current()
            qtype = "判断"
            continue

        question_match = re.match(r"^第\s*(\d+)\s*题\s*[,，、.．]?\s*(.*)$", text)
        if question_match:
            flush_current()
            current = {
                "level": level,
                "origin": source.name,
                "id": f"{level}-{source.stem}-{question_match.group(1)}",
                "type": qtype,
                "stem": clean(question_match.group(2)),
                "options": [],
                "answer": [],
            }
            continue

        if current is None:
            continue

        answer_match = re.match(r"^参考答案\s*[:：]\s*(.*)$", text)
        if answer_match:
            current["answer"] = normalize_answer(answer_match.group(1), str(current["type"]))
            continue

        if text.startswith("试题解析"):
            append_explanation(re.sub(r"^试题解析\s*[:：]?\s*", "", text))
            continue

        if current.get("answer"):
            append_explanation(text)
            continue

        option_match = re.match(r"^([A-J])\s*[、.．]\s*(.*)$", text)
        if option_match and current["type"] in {"单选", "多选"}:
            current["options"].append(  # type: ignore[union-attr]
                {"label": option_match.group(1), "text": clean(option_match.group(2))}
            )
            continue

        if current["type"] in {"单选", "多选"} and current["options"]:
            current["options"][-1]["text"] = clean(  # type: ignore[index,union-attr]
                str(current["options"][-1]["text"]) + " " + text  # type: ignore[index,union-attr]
            )
        else:
            current["stem"] = clean(str(current["stem"]) + " " + text)

    flush_current()
    return questions


NOISE_LINES = {
    "得分",
    "评分人",
    "姓名",
    "姓 名",
    "准考证号",
    "准 考 证 号",
    "地区",
    "地 区",
    "单位名称",
    "单 位 名 称",
}


def clean_pdf_line(line: str) -> str:
    text = line.replace("\uf0b7", "").replace("\uf0fc", "").strip()
    text = re.sub(r"^(地\s*区|准\s*考\s*证\s*号|姓\s*名)\s+", "", text).strip()
    return text


def denoise_pdf_block(block: str) -> str:
    lines = []
    for line in block.splitlines():
        text = clean_pdf_line(line)
        if not text:
            continue
        if compact(text) in {compact(item) for item in NOISE_LINES}:
            continue
        if "考生姓名" in text and "身份证号" in text and "准考证号" in text:
            continue
        lines.append(text)
    return "\n".join(lines)


def parse_pdf_options(text: str) -> tuple[str, list[dict[str, str]]]:
    matches = list(re.finditer(r"(?m)^\s*([A-J])、\s*(.*)$", text))
    if not matches:
        matches = list(re.finditer(r"(?m)^\s*([A-J])\.\s*(.*)$", text))
    if not matches:
        return text.strip(), []

    stem = text[: matches[0].start()].strip()
    options = []
    for index, match in enumerate(matches):
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        body = (match.group(2) + "\n" + text[match.end() : end]).strip()
        body = re.sub(r"\n+", " ", body).strip()
        if body:
            options.append({"label": match.group(1), "text": body})
    return stem, options


def pdf_level(source: Path) -> str:
    if "二级" in source.name:
        return "技师"
    if "三级" in source.name:
        return "高级工"
    if "四级" in source.name:
        return "中级工"
    raise ValueError(f"Cannot infer PDF level from {source.name}")


def source_label(source: Path) -> str:
    try:
        return str(source.relative_to(SOURCE_DIR))
    except ValueError:
        return source.name


def extract_pdf_bank(source: Path) -> list[dict[str, object]]:
    text = subprocess.check_output(["pdftotext", "-layout", str(source), "-"], text=True, errors="ignore")
    section_matches = list(re.finditer(r"[一二三]、\s*(单选题|多选题|判断题)", text))
    questions = []
    level = pdf_level(source)
    origin = source_label(source)
    type_map = {"单选题": "单选", "多选题": "多选", "判断题": "判断"}

    for section_index, section_match in enumerate(section_matches):
        qtype = type_map[section_match.group(1)]
        start = section_match.end()
        end = section_matches[section_index + 1].start() if section_index + 1 < len(section_matches) else len(text)
        section_text = text[start:end]
        answer_matches = list(
            re.finditer(r"【\s*参考答案\s*】\s*([A-J]+|正确|错误|对|错|√|×)", section_text)
        )
        previous_end = 0

        for answer_match in answer_matches:
            block = denoise_pdf_block(section_text[previous_end : answer_match.start()])
            previous_end = answer_match.end()
            number_matches = list(re.finditer(r"(?m)^\s*(\d+)、\s*", block))
            if not number_matches:
                number_matches = list(re.finditer(r"(?:^|\n).*?(\d+)、\s*", block))

            row_number = None
            if number_matches:
                question_match = number_matches[-1]
                row_number = int(question_match.group(1))
                question_text = block[question_match.end() :].strip()
            else:
                question_text = block.strip()

            if qtype in {"单选", "多选"}:
                stem, options = parse_pdf_options(question_text)
            else:
                stem, _pdf_options = parse_pdf_options(question_text)
                options = [{"label": "对", "text": "对"}, {"label": "错", "text": "错"}]

            stem = re.sub(r"\n+", " ", stem).strip()
            stem = re.sub(r"\s+", " ", stem)
            answer = normalize_answer(answer_match.group(1), qtype)
            if not stem or not answer:
                continue

            questions.append(
                {
                    "level": level,
                    "origin": origin,
                    "id": f"{level}-{source.parent.name}-{source.stem}-{row_number or len(questions) + 1}",
                    "type": qtype,
                    "stem": stem,
                    "options": options,
                    "answer": answer,
                }
            )

    return questions


def exact_key(question: dict[str, object]) -> tuple[object, ...]:
    options = tuple((option["label"], canonical(option["text"])) for option in question["options"])  # type: ignore[index]
    return (
        canonical(question["stem"]),
        question["type"],
        tuple(question["answer"]),  # type: ignore[arg-type]
        options,
    )


def extract() -> dict[str, object]:
    raw_questions = []
    source_counts: dict[str, int] = {}
    for level, source in SOURCES.items():
        bank_questions = extract_xlsx_bank(level, source)
        raw_questions.extend(bank_questions)
        source_counts[level] = len(bank_questions)

    docx_files: list[Path] = []
    for level, source in DOCX_SOURCES.items():
        bank_questions = extract_docx_bank(level, source)
        raw_questions.extend(bank_questions)
        source_counts[level] = source_counts.get(level, 0) + len(bank_questions)
        docx_files.append(source)

    pdf_counts: dict[str, int] = {}
    pdf_files = [source for source_dir in PDF_SOURCE_DIRS for source in sorted(source_dir.glob("*.pdf"))]
    for source in pdf_files:
        bank_questions = extract_pdf_bank(source)
        raw_questions.extend(bank_questions)
        level = pdf_level(source)
        source_counts[level] = source_counts.get(level, 0) + len(bank_questions)
        pdf_counts[level] = pdf_counts.get(level, 0) + len(bank_questions)

    merged: dict[tuple[object, ...], dict[str, object]] = {}
    for question in raw_questions:
        if question["type"] == "判断":
            question["options"] = [option.copy() for option in JUDGMENT_OPTIONS]
        key = exact_key(question)
        if key not in merged:
            merged[key] = {
                "stem": question["stem"],
                "type": question["type"],
                "options": question["options"],
                "answer": question["answer"],
                "levels": [],
                "origins": [],
                "explanations": [],
            }

        item = merged[key]
        level = str(question["level"])
        if level not in item["levels"]:  # type: ignore[operator]
            item["levels"].append(level)  # type: ignore[index,union-attr]
        origin = str(question.get("origin", ""))
        if origin and origin not in item["origins"]:  # type: ignore[operator]
            item["origins"].append(origin)  # type: ignore[index,union-attr]
        explanation = clean(question.get("explanation", ""))
        if explanation and explanation not in item["explanations"]:  # type: ignore[operator]
            item["explanations"].append(explanation)  # type: ignore[index,union-attr]

    level_order = {"中级工": 0, "高级工": 1, "技师": 2}
    questions: list[dict[str, object]] = []
    for index, item in enumerate(merged.values(), start=1):
        levels = sorted(item["levels"], key=lambda value: level_order[str(value)])  # type: ignore[arg-type]
        item["levels"] = levels
        item["scope"] = question_scope([str(level) for level in levels])
        item["id"] = f"Q{index:04d}"
        if not item["explanations"]:  # type: ignore[index]
            item.pop("explanations", None)
        questions.append(item)

    apply_explanation_overrides(questions)

    by_type: dict[str, int] = {}
    by_scope: dict[str, int] = {}
    for question in questions:
        qtype = str(question["type"])
        by_type[qtype] = by_type.get(qtype, 0) + 1
        scope = str(question["scope"])
        by_scope[scope] = by_scope.get(scope, 0) + 1

    return {
        "meta": {
            "title": "电力交易员中级工+高级工+技师题库",
            "sourceFiles": [source.name for source in SOURCES.values()]
            + [source.name for source in docx_files]
            + [f"{source_label(source_dir)}（PDF {len(list(source_dir.glob('*.pdf')))} 份）" for source_dir in PDF_SOURCE_DIRS],
            "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "rawTotal": len(raw_questions),
            "total": len(questions),
            "deduped": len(raw_questions) - len(questions),
            "sourceCounts": source_counts,
            "pdfSourceCounts": pdf_counts,
            "byType": by_type,
            "byScope": by_scope,
        },
        "questions": questions,
    }


def apply_explanation_overrides(questions: list[dict[str, object]]) -> None:
    if not EXPLANATION_OVERRIDES.exists():
        return

    overrides = json.loads(EXPLANATION_OVERRIDES.read_text(encoding="utf-8"))
    by_id = {str(question["id"]): question for question in questions}
    for question_id, entries in overrides.items():
        question = by_id.get(str(question_id))
        if not question or "高级工" not in question.get("levels", []):
            continue

        # Advanced-level-3 PDF explanations are maintained as a reviewed,
        # source-backed layer. Replace stale workbook/DOCX templates for
        # questions covered by the override file instead of showing both.
        explanations = []
        question["explanations"] = explanations
        for entry in entries:
            text = clean(entry.get("text", ""))
            source_title = clean(entry.get("sourceTitle", ""))
            source_url = clean(entry.get("sourceUrl", ""))
            confidence = clean(entry.get("confidence", ""))
            if not text:
                continue

            parts = [text]
            source_parts = []
            if source_title:
                source_parts.append(source_title)
            if source_url:
                source_parts.append(source_url)
            if source_parts:
                parts.append(f"来源：{' '.join(source_parts)}")
            if confidence:
                parts.append(f"可信度：{confidence}")
            explanation = "\n".join(parts)
            if explanation not in explanations:  # type: ignore[operator]
                explanations.append(explanation)  # type: ignore[union-attr]

def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    data = extract()
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    OUTPUT.write_text(f"window.QUESTION_BANK = {payload};\n", encoding="utf-8")
    write_split_files(data)
    update_index_script_tags(data)
    update_service_worker(data)
    meta = data["meta"]
    print(f"Wrote {OUTPUT}")
    print(f"Wrote split data chunks: {(len(data['questions']) + CHUNK_SIZE - 1) // CHUNK_SIZE}")
    print(f"Total: {meta['total']} | By type: {meta['byType']}")


def write_split_files(data: dict[str, object]) -> None:
    questions = data["questions"]
    meta = data["meta"]
    for old_chunk in DATA_DIR.glob("questions-*.js"):
        old_chunk.unlink()

    (DATA_DIR / "meta.js").write_text(
        "window.QUESTION_META = "
        + json.dumps(meta, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )
    for index in range(0, len(questions), CHUNK_SIZE):  # type: ignore[arg-type]
        part_no = index // CHUNK_SIZE + 1
        chunk = questions[index : index + CHUNK_SIZE]  # type: ignore[index]
        (DATA_DIR / f"questions-{part_no:02d}.js").write_text(
            "window.QUESTION_PARTS.push("
            + json.dumps(chunk, ensure_ascii=False, separators=(",", ":"))
            + ");\n",
            encoding="utf-8",
        )


def chunk_count(data: dict[str, object]) -> int:
    return (len(data["questions"]) + CHUNK_SIZE - 1) // CHUNK_SIZE  # type: ignore[arg-type]


def split_script_block(data: dict[str, object]) -> str:
    script_tags = [
        '  <script src="data/meta.js"></script>',
        '  <script>window.QUESTION_PARTS = [];</script>',
    ]
    for part_no in range(1, chunk_count(data) + 1):
        script_tags.append(f'  <script src="data/questions-{part_no:02d}.js"></script>')
    script_tags.append(
        '  <script>window.QUESTION_BANK = { meta: window.QUESTION_META, questions: window.QUESTION_PARTS.flat() };</script>'
    )
    return "\n".join(script_tags)


def update_index_script_tags(data: dict[str, object]) -> None:
    html = INDEX.read_text(encoding="utf-8")
    updated = SCRIPT_BLOCK_RE.sub("\n" + split_script_block(data), html)
    INDEX.write_text(updated, encoding="utf-8")


def update_service_worker(data: dict[str, object]) -> None:
    urls = [
        "./",
        "./index.html",
        "./manifest.webmanifest",
        "./icon.svg",
        "./quiz-core.js",
        "./data/meta.js",
        "./data/questions.js",
    ]
    urls.extend(f"./data/questions-{part_no:02d}.js" for part_no in range(1, chunk_count(data) + 1))
    block = "const APP_SHELL = [\n" + ",\n".join(f'  "{url}"' for url in urls) + "\n];"
    script = SERVICE_WORKER.read_text(encoding="utf-8")
    script = re.sub(r'const CACHE_NAME = ".*?";', 'const CACHE_NAME = "power-trader-quiz-v24";', script)
    script = APP_SHELL_RE.sub(block, script)
    SERVICE_WORKER.write_text(script, encoding="utf-8")


if __name__ == "__main__":
    main()
